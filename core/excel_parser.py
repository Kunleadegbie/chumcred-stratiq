# ==================================================
# core/excel_parser.py — Financial Excel Import (Strict)
# ==================================================

import openpyxl


# --------------------------------------------------
# TEMPLATE DEFINITION
# --------------------------------------------------

REQUIRED_SHEETS = {
    "Income_Statement": {
        "headers": ["Metric", "Y-2", "Y-1", "Y"],
        "metrics": ["Revenue", "EBITDA", "Net Profit"]
    },

    "Balance_Sheet": {
        "headers": ["Metric", "Value"],
        "metrics": [
            "Total Assets",
            "Equity",
            "Current Assets",
            "Current Liabilities",
            "Total Debt"
        ]
    },

    "Cash_Flow": {
        "headers": ["Metric", "Value"],
        "metrics": [
            "Operating Cash Flow",
            "CAPEX"
        ]
    }
}


# --------------------------------------------------
# PARSER
# --------------------------------------------------

def parse_financial_excel(file):

    try:
        wb = openpyxl.load_workbook(file, data_only=True)

    except Exception:
        raise ValueError("Invalid Excel file.")


    data = {}


    # Validate sheets
    for sheet, spec in REQUIRED_SHEETS.items():

        if sheet not in wb.sheetnames:
            raise ValueError(f"Missing sheet: {sheet}")


        ws = wb[sheet]


        # Validate headers
        headers = [c.value for c in ws[1]]

        if headers != spec["headers"]:
            raise ValueError(f"Invalid headers in {sheet}")


        values = {}


        for row in ws.iter_rows(min_row=2, values_only=True):

            metric = row[0]

            if metric not in spec["metrics"]:
                continue


            nums = row[1:]


            if any(v is None for v in nums):
                raise ValueError(f"Missing value in {sheet}: {metric}")


            values[metric] = [float(v) for v in nums]


        # Validate all metrics present
        for m in spec["metrics"]:

            if m not in values:
                raise ValueError(f"Missing metric: {m} in {sheet}")


        data[sheet] = values


    return data
